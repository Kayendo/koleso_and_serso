"""Экспорт истории игр игрока в XLSX."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from backend.models import PlayerGame, User
from backend.time_utils import ensure_aware

_STATUS_LABELS = {
    "completed": "Пройдена",
    "dropped": "Дроп",
    "active": "В процессе",
    "pending_admin": "Ждёт админа",
}

_HEADERS = [
    "Игра",
    "Клетка",
    "Жанр",
    "Кубик",
    "Статус",
    "Начало",
    "Завершение",
    "Время прохождения",
    "HLTB (ч)",
    "Оценка",
    "Отзыв",
    "Очки",
]


def _format_duration(seconds: int | None) -> str:
    if seconds is None:
        return ""
    s = max(0, int(seconds))
    h, m, sec = s // 3600, (s % 3600) // 60, s % 60
    return f"{h} ч {m} м {sec} с"


def _as_excel_dt(value):
    if not value:
        return None
    return ensure_aware(value).replace(tzinfo=None)


def build_games_xlsx(user: User) -> BytesIO:
    games = (
        PlayerGame.query.filter_by(user_id=user.id)
        .order_by(PlayerGame.created_at.asc(), PlayerGame.id.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Игры"

    ws.append([f"Игрок: {user.public_name()}"])
    ws.append([f"Всего игр: {len(games)}"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(_HEADERS)
    for col in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)

    for game in games:
        ws.append(
            [
                game.title,
                game.cell_name or "",
                game.genre_label or "",
                game.dice_roll or "",
                _STATUS_LABELS.get(game.status or "", game.status or ""),
                _as_excel_dt(game.created_at),
                _as_excel_dt(game.finished_at),
                _format_duration(game.play_seconds),
                game.hltb_hours,
                game.rating,
                (game.review or "").strip(),
                game.points_earned,
            ]
        )

    for col in range(1, len(_HEADERS) + 1):
        letter = get_column_letter(col)
        max_len = len(_HEADERS[col - 1])
        for row in ws.iter_rows(
            min_row=header_row + 1, max_row=ws.max_row, min_col=col, max_col=col
        ):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, min(len(str(val)), 80))
        ws.column_dimensions[letter].width = min(max_len + 2, 48)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def safe_export_filename(username: str) -> str:
    safe = "".join(c if c.isalnum() or c in "._- " else "_" for c in username)
    safe = safe.strip().replace(" ", "_") or "player"
    return f"{safe}_games.xlsx"

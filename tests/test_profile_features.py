"""Смена отображаемого имени и экспорт игр в XLSX."""

from io import BytesIO

from openpyxl import load_workbook

from backend.accounts import PLAYER_ACCOUNTS
from backend.models import PlayerGame, User, db
from backend.services.turn_service import create_player_game

from tests.conftest import reset_player


def test_rename_display_name_keeps_login(app, player_client, actor):
    with app.app_context():
        reset_player(actor)
        login_name = PLAYER_ACCOUNTS[0]["username"]
        actor.username = login_name
        actor.display_name = login_name
        db.session.commit()
        user_id = actor.id

    r = player_client.patch(
        "/api/me/display-name", json={"displayName": "Красивое Имя"}
    )
    assert r.status_code == 200
    data = r.get_json()["user"]
    assert data["displayName"] == "Красивое Имя"
    assert data["username"] == login_name

    with app.app_context():
        u = User.query.get(user_id)
        assert u.username == login_name
        assert u.display_name == "Красивое Имя"

    player_client.patch(
        "/api/me/display-name", json={"displayName": login_name}
    )


def test_rename_duplicate_rejected(app, player_client, actor, second_player):
    with app.app_context():
        reset_player(actor)
        other_name = second_player.public_name()

    r = player_client.patch(
        "/api/me/display-name", json={"displayName": other_name}
    )
    assert r.status_code == 400


def test_export_games_xlsx(app, client, actor):
    with app.app_context():
        user = User.query.get(actor.id)
        reset_player(user)
        user.display_name = "Exporter"
        db.session.commit()
        create_player_game(user, "Export Game", 5, "3+4")
        game = PlayerGame.query.filter_by(user_id=user.id).first()
        game.status = "completed"
        game.review = "Круто"
        game.rating = 9
        game.play_seconds = 3661
        game.points_earned = 5
        db.session.commit()
        user_id = user.id
        public_name = user.public_name()

    assert public_name == "Exporter"

    resp = client.get(f"/api/players/{user_id}/games.xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.mimetype
    wb = load_workbook(BytesIO(resp.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    flat = [str(c) for row in rows for c in row if c is not None]
    assert any(public_name in s for s in flat)
    assert any("Export Game" in s for s in flat)
    assert any("Круто" in s for s in flat)
    assert any("1 ч 1 м 1 с" in s for s in flat)
